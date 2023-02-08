# -*- coding: utf-8 -*-
"""
Created on Thu Feb  2 15:03:47 2023

@author: Weave
"""
import sys
from pathlib import Path
py_dir = Path(__file__).parent

import openpyxl
import math
import itertools
import copy

import matplotlib.pyplot as plt
import matplotlib.colors as mcolors

from loguru import logger
from matplotlib.patches import Rectangle

pop_gen = lambda gen, i0: list(itertools.islice(gen, i0+1))[-1]

def plot_colortable(colors, *, ncols=4, sort_colors=True):
    """
    展示颜色所用的，可以忽略
    
    plot_colortable(mcolors.CSS4_COLORS)
    plt.show()
    
    Parameters
    ----------
    colors : TYPE
        DESCRIPTION.
    * : TYPE
        DESCRIPTION.
    ncols : TYPE, optional
        DESCRIPTION. The default is 4.
    sort_colors : TYPE, optional
        DESCRIPTION. The default is True.

    Returns
    -------
    fig : TYPE
        DESCRIPTION.

    """
    cell_width = 212
    cell_height = 22
    swatch_width = 48
    margin = 12

    # Sort colors by hue, saturation, value and name.
    if sort_colors is True:
        names = sorted(
            colors, key=lambda c: tuple(mcolors.rgb_to_hsv(mcolors.to_rgb(c))))
    else:
        names = list(colors)

    n = len(names)
    nrows = math.ceil(n / ncols)

    width = cell_width * 4 + 2 * margin
    height = cell_height * nrows + 2 * margin
    dpi = 72

    fig, ax = plt.subplots(figsize=(width / dpi, height / dpi), dpi=dpi)
    fig.subplots_adjust(margin/width, margin/height,
                        (width-margin)/width, (height-margin)/height)
    ax.set_xlim(0, cell_width * 4)
    ax.set_ylim(cell_height * (nrows-0.5), -cell_height/2.)
    ax.yaxis.set_visible(False)
    ax.xaxis.set_visible(False)
    ax.set_axis_off()

    for i, name in enumerate(names):
        row = i % nrows
        col = i // nrows
        y = row * cell_height

        swatch_start_x = cell_width * col
        text_pos_x = cell_width * col + swatch_width + 7

        ax.text(text_pos_x, y, name, fontsize=14,
                horizontalalignment='left',
                verticalalignment='center')

        ax.add_patch(
            Rectangle(xy=(swatch_start_x, y-9), width=swatch_width,
                      height=18, facecolor=colors[name], edgecolor='0.7')
        )

    return fig

class ExcelStyle:
    
    def __init__(self, fp: Path, sheet_name=0):
        """

        Parameters
        ----------
        fp : Path
            Excel文件的存储路径.
        sheet_name : TYPE, optional
            需要调整的sheet名称. The default is 0.
        """
        self.fp = fp
        self.sheet_name = sheet_name
        if not self.fp.exists() or self.fp.is_dir():
            raise Exception(f"{str(fp)}文件不存在或不是文件")
        
    def __enter__(self):
        """
        使用with的时候调用
        with ExcelStyle(Path, "format_stat") as obj:
            # obj.cell_solid_fill(2, "B", "orange")
            # obj.cell_solid_fill((3,2), "orange")
            # obj.map_col(1, obj.cell_solid_fill, args=["orange"])
            obj.col_auto_width()
        """
        self.wb = openpyxl.load_workbook(self.fp)
        if self.sheet_name not in self.wb.sheetnames:
            raise Exception(f"Sheet{self.sheetnames}不存在")
        
        self.ws = self.wb[self.sheet_name]
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.wb.save(self.fp)
        logger.info(f"Successfully saved: {self.fp}")
        self.wb.close()
        
    def get_cell_obj(self, *args):
        """

        Parameters
        ----------
        *args : TYPE
            只有一个传参时，先判断是不是列表或者元祖等对象，若是，则用此对象替换传参args。
            若传入的一个对象本来就是单元格对象，那么直接返回.
            若传入的是一个字符串，则认为是"A1"这种单元格表达式，返回对应的单元格对象.
            若传入的是多个参数，则认为第一个是行号，第二个是列名，返回对应的单元格对象.
        """
        if len(args) == 1 and hasattr(args[0], '__class_getitem__'):
            args = args[0]
        
        # 根据不同的传参情况，找到对应的单元格对象
        if len(args) == 1 and isinstance(args[0], openpyxl.cell.cell.Cell):
            return args[0] # 传入的一个对象本来就是单元格对象
        # 传入的是一个字符串
        elif len(args) == 1 and isinstance(args[0], str):
            return self.ws[args[0]]
        # 传入的是多个参数
        elif len(args) == 2:
            row_id = int(args[0])
            col_id = args[1]
            if isinstance(col_id, str) and col_id.isalpha(): # 若传入的是字母，则替换成数字列名
                col_id = openpyxl.utils.column_index_from_string(col_id)
            return self.ws.cell(row_id, col_id)
        else:
            raise Exception(f"参数有问题：{str(args)}")
            
    def col_map(self, col_id, func, args=[]):
        """
        使用传入的函数遍历列
        obj.col_map("A", obj.cell_solid_fill, args=["orange"])
        
        Parameters
        ----------
        col_id : TYPE
            列名.
        func : TYPE
            函数.
        args : TYPE, optional
            传入的函数中，所需的参数. The default is [].
        """
        # 
        if isinstance(col_id, str) and col_id.isalpha():
            col_id = openpyxl.utils.column_index_from_string(col_id)
        else:
            col_id = int(col_id)
        
        col_obj = pop_gen(self.ws.columns, col_id-1)
        for c0 in col_obj:
            func(c0, *args)
            
    def row_map(self, row_id, func, args=[]):
        """
        使用传入的函数遍历行
        obj.row_map(1, obj.cell_solid_fill, args=["orange"])
        
        Parameters
        ----------
        col_id : TYPE
            行名.
        func : TYPE
            函数.
        args : TYPE, optional
            传入的函数中，所需的参数. The default is [].
        """
        row_id = int(row_id)
        
        row_obj = pop_gen(self.ws.rows, row_id-1)
        for c0 in row_obj:
            func(c0, *args)
                    
    def cell_solid_fill(self, cell, color, mcolor=mcolors.CSS4_COLORS):
        """
        将指定cell的背景颜色填充为指定颜色
        """
        if not color.startswith("#"):
            color = mcolor[color]
            
        cell_obj = self.get_cell_obj(cell)
        patt = openpyxl.styles.PatternFill('solid', fgColor=color[1:])
        cell_obj.fill = patt
        
    def cols_auto_width(self):
        for col in self.ws.columns:
            # 获取列序号
            index = list(self.ws.columns).index(col)
            # 获取行字母表头
            letter = openpyxl.utils.get_column_letter(index+1)
            # 获取当前列最大宽度
            col_len = max(map(lambda x:len(str(x.value)), col))
            # 设置列宽为最大长度比例
            self.ws.column_dimensions[letter].width = col_len*1.1

if __name__ == "__main__":
    # plot_colortable(mcolors.CSS4_COLORS)
    # plt.show()
    # print(mcolors.CSS4_COLORS)
    with ExcelStyle(Path(r"C:\Users\Weave\Desktop\易查云\资源库\ktr文件管理\tmp\source_report20230207.xlsx"), "format_stat") as obj:
        # obj.cell_solid_fill(2, "B", "orange")
        # obj.cell_solid_fill((3,2), "orange")
        # obj.map_col(1, obj.cell_solid_fill, args=["orange"])
        obj.col_auto_width()
